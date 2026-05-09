import sys
from collections import deque
import pandas as pd
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTextEdit, QLabel, QComboBox, QMessageBox,
    QTableWidget, QTableWidgetItem, QCheckBox, QDialog, QDialogButtonBox,
    QListWidget, QListWidgetItem, QLineEdit, QFormLayout, QPlainTextEdit,
    QFileDialog, QInputDialog, QProgressBar
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSettings
from PyQt6.QtGui import QFont, QTextDocument, QTextCursor, QColor, QAction, QIcon
from tr import ExcelParser
from _version import __version__
from templates import (
    TemplateRegistry,
    Template,
    HighlightRule,
    USER_TEMPLATES_PATH,
)
from view import ViewMeta, apply_template, parse_location_parts


import json as _json


# ---------------------------------------------------------------------------
# Edit history (undo / redo)
# ---------------------------------------------------------------------------

UNDO_HISTORY_LIMIT = 10  # number of undoable steps to keep in memory


class _CheckboxEdit:
    """Records one user checkbox click (may affect multiple rows via location grouping)."""
    __slots__ = ("affected",)

    def __init__(self, affected):
        # [(table_row_idx, old_checked, new_checked), ...]
        self.affected = affected


class _CellTextEdit:
    """Records a single cell-text change made by the user."""
    __slots__ = ("row", "col", "old_text", "new_text")

    def __init__(self, row: int, col: int, old_text: str, new_text: str):
        self.row = row
        self.col = col
        self.old_text = old_text
        self.new_text = new_text


class _EditHistory:
    """Fixed-size undo / redo stack for table edits."""

    def __init__(self, limit: int = UNDO_HISTORY_LIMIT) -> None:
        self._undo: deque = deque(maxlen=limit)
        self._redo: deque = deque()

    def push(self, edit) -> None:
        """Record a new edit; clears the redo stack."""
        self._undo.append(edit)
        self._redo.clear()

    def undo(self):
        if not self._undo:
            return None
        edit = self._undo.pop()
        self._redo.append(edit)
        return edit

    def redo(self):
        if not self._redo:
            return None
        edit = self._redo.pop()
        self._undo.append(edit)
        return edit

    def can_undo(self) -> bool:
        return bool(self._undo)

    def can_redo(self) -> bool:
        return bool(self._redo)

    def clear(self) -> None:
        self._undo.clear()
        self._redo.clear()


class TemplatesDialog(QDialog):
    """Minimal CRUD dialog over the user template store.

    Editing happens as raw JSON for the selected template - this keeps the
    dialog small while still exposing every option in :class:`templates.Template`
    (drop, rename, order, sort_by, location_columns, highlights, etc.). Wiring
    up form widgets per field can come later without changing the storage
    format.
    """

    def __init__(self, registry: TemplateRegistry, parent=None):
        super().__init__(parent)
        self.registry = registry
        self.setWindowTitle("DocuReader Templates")
        self.resize(900, 600)
        self._dirty = False

        layout = QHBoxLayout(self)

        # Left: list of templates.
        left = QVBoxLayout()
        left.addWidget(QLabel("Templates"))
        self.list_widget = QListWidget()
        self.list_widget.currentRowChanged.connect(self._on_select)
        left.addWidget(self.list_widget)

        list_buttons = QHBoxLayout()
        self.new_button = QPushButton("New")
        self.new_button.clicked.connect(self._on_new)
        list_buttons.addWidget(self.new_button)
        self.delete_button = QPushButton("Delete")
        self.delete_button.clicked.connect(self._on_delete)
        list_buttons.addWidget(self.delete_button)
        left.addLayout(list_buttons)

        io_buttons = QHBoxLayout()
        self.import_button = QPushButton("Import...")
        self.import_button.clicked.connect(self._on_import)
        io_buttons.addWidget(self.import_button)
        self.export_button = QPushButton("Export...")
        self.export_button.clicked.connect(self._on_export)
        io_buttons.addWidget(self.export_button)
        left.addLayout(io_buttons)

        layout.addLayout(left, 1)

        # Right: JSON editor for the selected template.
        right = QVBoxLayout()
        right.addWidget(QLabel(f"Editing template (stored at {USER_TEMPLATES_PATH})"))
        self.editor = QPlainTextEdit()
        self.editor.setFont(QFont("Courier New", 10))
        self.editor.textChanged.connect(self._mark_dirty)
        right.addWidget(self.editor)

        self.error_label = QLabel("")
        self.error_label.setStyleSheet("QLabel { color: #b00; }")
        right.addWidget(self.error_label)

        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save
            | QDialogButtonBox.StandardButton.Close
        )
        button_box.button(QDialogButtonBox.StandardButton.Save).clicked.connect(self._on_save)
        button_box.button(QDialogButtonBox.StandardButton.Close).clicked.connect(self.accept)
        right.addWidget(button_box)

        layout.addLayout(right, 2)

        self._reload_list()

    def _reload_list(self):
        self.list_widget.clear()
        for t in self.registry.templates:
            label = t.name + ("  (built-in)" if t.builtin else "")
            self.list_widget.addItem(QListWidgetItem(label))
        if self.registry.templates:
            self.list_widget.setCurrentRow(0)

    def _current_template(self) -> Optional[Template]:
        row = self.list_widget.currentRow()
        if row < 0 or row >= len(self.registry.templates):
            return None
        return self.registry.templates[row]

    def _on_select(self, _row: int):
        t = self._current_template()
        if t is None:
            self.editor.setPlainText("")
            return
        self.editor.blockSignals(True)
        self.editor.setPlainText(_json.dumps(t.to_dict(), indent=2))
        self.editor.blockSignals(False)
        self._dirty = False
        self.error_label.setText("")

    def _mark_dirty(self):
        self._dirty = True
        self.error_label.setText("Unsaved changes.")

    def _parse_editor(self) -> Optional[Template]:
        try:
            data = _json.loads(self.editor.toPlainText())
        except _json.JSONDecodeError as e:
            self.error_label.setText(f"Invalid JSON: {e}")
            return None
        try:
            return Template.from_dict(data)
        except (KeyError, TypeError, ValueError) as e:
            self.error_label.setText(f"Invalid template: {e}")
            return None

    def _on_save(self):
        t = self._parse_editor()
        if t is None:
            return
        original = self._current_template()
        if original is not None and original.name != t.name:
            self.registry.remove(original.name)
        self.registry.upsert(t)
        try:
            self.registry.save()
        except OSError as e:
            self.error_label.setText(f"Could not write user templates: {e}")
            return
        self._dirty = False
        self.error_label.setText("Saved.")
        current_row = self.list_widget.currentRow()
        self._reload_list()
        if current_row >= 0:
            self.list_widget.setCurrentRow(min(current_row, self.list_widget.count() - 1))

    def _on_new(self):
        new = Template(
            name=f"New Template {len(self.registry.templates) + 1}",
            description="Describe when this template applies.",
            filename_patterns=["*example*"],
            required_columns=[],
            drop=[],
            order=[],
            sort_by=[],
            location_columns=[],
            highlights=[],
        )
        self.registry.upsert(new)
        try:
            self.registry.save()
        except OSError as e:
            self.error_label.setText(f"Could not write user templates: {e}")
        self._reload_list()
        self.list_widget.setCurrentRow(len(self.registry.templates) - 1)

    def _on_delete(self):
        t = self._current_template()
        if t is None:
            return
        if t.builtin:
            QMessageBox.information(
                self,
                "Built-in Template",
                "Built-in templates cannot be deleted. Edit and rename to create a copy.",
            )
            return
        if QMessageBox.question(
            self,
            "Delete Template",
            f"Delete template '{t.name}'? This cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return
        self.registry.remove(t.name)
        try:
            self.registry.save()
        except OSError as e:
            self.error_label.setText(f"Could not write user templates: {e}")
        self._reload_list()

    def _on_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import template", "", "JSON (*.json)")
        if not path:
            return
        try:
            data = _json.loads(Path(path).read_text(encoding="utf-8"))
            t = Template.from_dict(data)
        except (OSError, _json.JSONDecodeError, KeyError, TypeError, ValueError) as e:
            QMessageBox.critical(self, "Import failed", str(e))
            return
        self.registry.upsert(t)
        try:
            self.registry.save()
        except OSError as e:
            QMessageBox.critical(self, "Import failed", str(e))
            return
        self._reload_list()

    def _on_export(self):
        t = self._current_template()
        if t is None:
            return
        path, _ = QFileDialog.getSaveFileName(self, "Export template", f"{t.name}.json", "JSON (*.json)")
        if not path:
            return
        try:
            Path(path).write_text(_json.dumps(t.to_dict(), indent=2), encoding="utf-8")
        except OSError as e:
            QMessageBox.critical(self, "Export failed", str(e))


# ---------------------------------------------------------------------------
# Update helpers — download thread + progress dialog
# ---------------------------------------------------------------------------

class _UpdateThread(QThread):
    """Downloads and stages a DocuReader release without blocking the GUI.

    Signals
    -------
    progress(downloaded, total)
        Emitted during the download with cumulative byte counts.
    status(message)
        Human-readable status string suitable for display in the dialog.
    staged(staged_dir, install_dir)
        Emitted once the release has been downloaded, verified, and extracted.
        Both values are :class:`pathlib.Path` objects.
    error(message)
        Emitted on any failure (network, checksum, cancellation-free error).
    """

    progress = pyqtSignal(int, int)
    status = pyqtSignal(str)
    staged = pyqtSignal(object, object)
    error = pyqtSignal(str)

    def __init__(self, include_prereleases: bool = False) -> None:
        super().__init__()
        self.include_prereleases = include_prereleases
        self._cancelled = False

    def cancel(self) -> None:
        """Signal the thread to stop at the next progress callback."""
        self._cancelled = True

    def run(self) -> None:
        import updater_github

        try:
            self.status.emit("Checking for latest release…")
            release = updater_github.fetch_release(self.include_prereleases)
            if release is None:
                self.error.emit(
                    "Could not reach GitHub Releases.\n"
                    "Check your internet connection and try again."
                )
                return

            if not updater_github.is_newer(release.tag, updater_github.CURRENT_VERSION):
                self.error.emit(
                    f"Already on the latest version ({updater_github.CURRENT_VERSION})."
                )
                return

            self.status.emit(f"Downloading {release.tag}…")

            def _on_progress(downloaded: int, total: int) -> None:
                if self._cancelled:
                    raise RuntimeError("Update cancelled by user.")
                self.progress.emit(downloaded, total)

            staged = updater_github.stage_release(release, progress=_on_progress)
            if staged is None:
                if not self._cancelled:
                    self.error.emit(
                        "Download failed or checksum mismatch.\n"
                        "The update was not applied."
                    )
                return

            install_dir = updater_github.install_dir_for_running_exe()
            self.staged.emit(staged, install_dir)

        except Exception as e:
            if not self._cancelled:
                self.error.emit(str(e))


class _UpdateProgressDialog(QDialog):
    """Modal progress dialog shown while a release is being downloaded."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Downloading Update")
        self.setModal(True)
        self.setMinimumWidth(440)
        # Disable the X close button so the user must use Cancel.
        self.setWindowFlags(
            self.windowFlags()
            & ~Qt.WindowType.WindowContextHelpButtonHint
            & ~Qt.WindowType.WindowCloseButtonHint
        )

        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        self.status_label = QLabel("Initializing…")
        layout.addWidget(self.status_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)

        self.detail_label = QLabel("")
        self.detail_label.setStyleSheet("QLabel { color: #666; font-size: 9pt; }")
        layout.addWidget(self.detail_label)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        btn_row.addWidget(self.cancel_button)
        layout.addLayout(btn_row)

    def set_status(self, text: str) -> None:
        self.status_label.setText(text)

    def set_progress(self, downloaded: int, total: int) -> None:
        if total > 0:
            pct = int(downloaded * 100 / total)
            self.progress_bar.setValue(pct)
            mb_down = downloaded / (1024 * 1024)
            mb_total = total / (1024 * 1024)
            self.detail_label.setText(f"{mb_down:.1f} MB / {mb_total:.1f} MB")
        else:
            self.progress_bar.setRange(0, 0)  # indeterminate spinner


class WorkerThread(QThread):
    """Worker thread to run the parser without blocking the GUI"""
    finished = pyqtSignal()
    error = pyqtSignal(str)
    output = pyqtSignal(str)
    table_ready = pyqtSignal(pd.DataFrame, object)  # df, ViewMeta
    data_loaded = pyqtSignal(object)  # Emit the dataframe
    task_ids_ready = pyqtSignal(list)  # Emit the task IDs for clipboard copy
    template_matched = pyqtSignal(str, str)  # template_name, match_reason
    
    def __init__(self, filepaths: List[str], function_name: str, registry: TemplateRegistry, sheet_names: Optional[Dict[str, str]] = None):
        super().__init__()
        self.filepaths = filepaths
        self.filepath = filepaths[0] if filepaths else ""  # primary path for template matching
        self.function_name = function_name
        self.registry = registry
        self.sheet_names = sheet_names or {}  # filepath -> sheet_name
        self.parser = None
        self.matched_template: Optional[Template] = None

    def format_df(self, df: pd.DataFrame) -> str:
        """Format a DataFrame into an aligned text table for the log pane."""
        if df is None or df.empty:
            return "No data to display\n"

        cols = list(df.columns)
        rows: List[List[str]] = []
        col_widths: Dict[str, int] = {c: len(str(c)) for c in cols}

        for _, r in df.iterrows():
            row = ["" if pd.isna(r[c]) else str(r[c]) for c in cols]
            rows.append(row)
            for i, val in enumerate(row):
                col_name = cols[i]
                if len(val) > col_widths[col_name]:
                    col_widths[col_name] = len(val)

        header_parts = [str(c).ljust(col_widths[c]) for c in cols]
        sep_parts = ["-" * col_widths[c] for c in cols]
        header = " | ".join(header_parts)
        separator = "-+-".join(sep_parts)

        lines = [header, separator]
        for row in rows:
            row_parts = [row[i].ljust(col_widths[cols[i]]) for i in range(len(cols))]
            lines.append(" | ".join(row_parts))

        return "\n".join(lines) + "\n"

    def run(self):
        try:
            all_dfs: List[pd.DataFrame] = []
            for fp in self.filepaths:
                parser = ExcelParser(fp)
                sheet = self.sheet_names.get(fp)
                if sheet:
                    parser.read_excel(sheet_name=sheet)
                else:
                    parser.read_excel()
                if parser.df is None:
                    self.error.emit(f"Failed to load {Path(fp).name}")
                    return
                self.output.emit(f"Loaded '{Path(fp).name}'\n")
                all_dfs.append(parser.df)

            if len(all_dfs) == 1:
                combined_df = all_dfs[0]
            else:
                combined_df = pd.concat(all_dfs, ignore_index=True)
                self.output.emit(
                    f"Combined {len(self.filepaths)} files into one dataset\n"
                )

            # Match template using the primary file's name and the combined columns.
            match = self.registry.select(combined_df.columns, self.filepath)
            self.matched_template = match.template
            self.output.emit(f"Detected template: {match.template.name}\n")
            self.template_matched.emit(match.template.name, match.reason)

            # Build a single parser whose .df is the combined data so that
            # execute_selected_function() works unchanged.
            self.parser = ExcelParser(self.filepath)
            self.parser.df = combined_df

            self.execute_selected_function()
            self.data_loaded.emit(self.parser.df)

        except Exception as e:
            self.error.emit(f"Error: {str(e)}")
        finally:
            self.finished.emit()
    
    def execute_selected_function(self):
        """Execute the selected function from the dropdown."""
        try:
            if self.parser is None or self.parser.df is None:
                self.output.emit("Parser is not initialized.\n")
                return

            template = self.matched_template
            template_name = template.name if template else ""

            if self.function_name == "get_task_ids_where_condition":
                # Special-case the Locked Full Container template (was an inline
                # column-set sniff in tr.py / tr_gui.py; now driven by template name).
                if template_name == "Locked Full Container Chase Tasks":
                    task_ids = self.parser.get_unique_numeric_values("TASK_ID")
                    self.output.emit(f"Task IDs found: {task_ids}\n")
                    items_df = pd.DataFrame(columns=["Item", "Affected Task ID Count"])
                    self.table_ready.emit(items_df, ViewMeta(template_name=template_name))
                    self.task_ids_ready.emit(task_ids)
                    return

                task_ids, items_not_met = self.parser.get_task_ids_where_condition(
                    task_id_col="Task ID",
                    condition_col1="Active OHB",
                    condition_col2="Allocated",
                    comparison=">=",
                    item_col="Item",
                )

                self.output.emit("Items still needing replenishment:\n")
                if items_not_met:
                    for item, task_id_count in sorted(items_not_met.items(), key=lambda x: x[1], reverse=True):
                        self.output.emit(f"  {item}: {task_id_count} task(s)\n")
                else:
                    self.output.emit("  None — all items are stocked.\n")

                self.output.emit(f"Tasks ready to release: {task_ids}\n")

                if items_not_met:
                    sorted_items = sorted(items_not_met.items(), key=lambda x: x[1], reverse=True)
                    items_df = pd.DataFrame(sorted_items, columns=["Item", "Affected Task ID Count"])
                else:
                    items_df = pd.DataFrame(columns=["Item", "Affected Task ID Count"])

                self.table_ready.emit(items_df, ViewMeta(template_name=template_name))
                self.task_ids_ready.emit(task_ids)

            elif self.function_name == "display_all":
                if template is None:
                    self.output.emit("No template matched — showing raw data.\n")
                    self.table_ready.emit(self.parser.df.reset_index(drop=True), ViewMeta())
                    return
                prepared_df, meta = apply_template(self.parser.df, template)
                self.table_ready.emit(prepared_df, meta)

            else:
                self.output.emit(f"Unknown function: {self.function_name}\n")

        except Exception as e:
            self.output.emit(f"Error executing function: {str(e)}\n")


class ExcelParserGUI(QMainWindow):
    """PyQt GUI for the Excel Parser"""
    
    def __init__(self):
        super().__init__()
        self.worker_thread = None
        self.current_df = None
        self.task_ids = None
        self.strikethrough_rows = set()  # Track which rows have strikethrough applied
        self.cell_highlights: Dict[Tuple[int, str], str] = {}  # (data_row, col) -> color
        self.table_row_map = None  # Map table row index to data row index
        self.table_row_location_values = None  # Map table row index to location value
        self.location_col = None  # Track location column for table grouping
        self.bulk_checkbox_update = False  # Prevent recursive checkbox handling
        self._update_thread: Optional[_UpdateThread] = None
        self.registry = TemplateRegistry.load()
        self.settings = QSettings("DocuReader", "DocuReader")
        # Track previous version for the "Version Info" dialog.
        stored = self.settings.value("version/current", "", type=str)
        if stored and stored != __version__:
            self.settings.setValue("version/previous", stored)
        self.settings.setValue("version/current", __version__)
        self.view_df: Optional[pd.DataFrame] = None
        self.view_meta: Optional[ViewMeta] = None
        self._edit_history = _EditHistory()
        self._suppressing_history = False
        self._pre_edit_text: Dict[Tuple[int, int], str] = {}
        self.init_ui()
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("Lomar Inventory Control - DocuReader")
        _icon_path = (
            Path(sys.executable).parent / "DocuReader.ico"
            if getattr(sys, "frozen", False)
            else Path(__file__).parent / "DocuReader.ico"
        )
        if _icon_path.exists():
            self.setWindowIcon(QIcon(str(_icon_path)))
        screen = QApplication.primaryScreen().availableGeometry()
        w = int(screen.width() * 0.58)
        h = int(screen.height() * 0.88)
        self.resize(w, h)

        # ----- Menu bar -----
        tools_menu = self.menuBar().addMenu("&Tools")

        self.update_action = QAction("Check && Install Updates", self)
        self.update_action.triggered.connect(self.check_for_updates)
        tools_menu.addAction(self.update_action)

        self.prerelease_action = QAction("Include Pre-Releases", self)
        self.prerelease_action.setCheckable(True)
        self.prerelease_action.setChecked(
            self.settings.value("updater/include_prereleases", False, type=bool)
        )
        self.prerelease_action.toggled.connect(
            lambda v: self.settings.setValue("updater/include_prereleases", bool(v))
        )
        tools_menu.addAction(self.prerelease_action)

        tools_menu.addSeparator()

        templates_action = QAction("Templates...", self)
        templates_action.triggered.connect(self.open_templates_dialog)
        tools_menu.addAction(templates_action)

        tools_menu.addSeparator()

        self.export_action = QAction("Export View...", self)
        self.export_action.triggered.connect(self.export_view)
        self.export_action.setEnabled(False)
        tools_menu.addAction(self.export_action)

        batch_action = QAction("Batch Export...", self)
        batch_action.triggered.connect(self.batch_export)
        tools_menu.addAction(batch_action)

        tools_menu.addSeparator()

        version_action = QAction("Version Info...", self)
        version_action.triggered.connect(self.show_version_info)
        tools_menu.addAction(version_action)

        tools_menu.addSeparator()

        terminate_action = QAction("Terminate", self)
        terminate_action.triggered.connect(self.terminate_program)
        tools_menu.addAction(terminate_action)

        # Create central widget
        central_widget = QWidget()
        central_widget.setObjectName("central_widget")
        self.setCentralWidget(central_widget)
        
        # Create main layout
        main_layout = QVBoxLayout()
        
        # Title
        title = QLabel("Inventory DocuReader")
        title_font = title.font()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        self.title_label = title
        main_layout.addWidget(title)
        
        # Status banner — centered at the top, shows user-relevant activity messages.
        self.status_label = QLabel("Ready")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        _app_palette = QApplication.instance().palette()
        _is_dark = _app_palette.color(_app_palette.ColorRole.Window).lightness() < 128
        _status_color = "#ffffff" if _is_dark else "#000000"
        self.status_label.setStyleSheet(
            f"QLabel {{ font-style: italic; font-weight: bold; color: {_status_color}; padding: 3px 0; }}"
        )
        main_layout.addWidget(self.status_label)
        
        # File selection section
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Which file do you want?"))
        
        self.file_combo = QComboBox()
        self.populate_downloads_files()
        file_layout.addWidget(self.file_combo)
        
        main_layout.addLayout(file_layout)

        # Extra files row — allows multi-document parsing (e.g. east + west tower).
        extra_layout = QHBoxLayout()
        self.extra_files_label = QLabel("Additional files: none")
        self.extra_files_label.setStyleSheet("QLabel { color: #555; font-style: italic; }")
        extra_layout.addWidget(self.extra_files_label)
        self.add_files_button = QPushButton("Add files...")
        self.add_files_button.setToolTip(
            "Select additional CSV/XLSX files to parse together with the primary file above. "
            "Rows from all files are combined before analysis."
        )
        self.add_files_button.clicked.connect(self.add_extra_files)
        extra_layout.addWidget(self.add_files_button)
        self.remove_extra_file_button = QPushButton("Remove selected")
        self.remove_extra_file_button.clicked.connect(self.remove_extra_file)
        extra_layout.addWidget(self.remove_extra_file_button)
        self.clear_extra_button = QPushButton("Clear all")
        self.clear_extra_button.clicked.connect(self.clear_extra_files)
        extra_layout.addWidget(self.clear_extra_button)
        extra_layout.addStretch()
        main_layout.addLayout(extra_layout)

        self.extra_files_list = QListWidget()
        self.extra_files_list.setMaximumHeight(70)
        self.extra_files_list.hide()
        main_layout.addWidget(self.extra_files_list)

        # Function selection section
        function_layout = QHBoxLayout()
        function_layout.addWidget(QLabel("What do you need?"))
        
        self.function_combo = QComboBox()
        self.function_combo.addItem("Chase Tasks Needing Released", "get_task_ids_where_condition")
        #self.function_combo.addItem("filter_by_value", "filter_by_value")
        #self.function_combo.addItem("filter_by_range", "filter_by_range")
        #self.function_combo.addItem("filter_by_contains", "filter_by_contains")
        self.function_combo.addItem("The Whole Table", "display_all")
        
        function_layout.addWidget(self.function_combo)
        main_layout.addLayout(function_layout)

        # Detected template / category label (populated after analysis runs).
        self.detected_label = QLabel("Detected category: (none yet - run an analysis)")
        self.detected_label.setStyleSheet("QLabel { color: #555; font-style: italic; }")
        main_layout.addWidget(self.detected_label)

        # Buttons layout
        button_layout = QHBoxLayout()
        
        self.start_button = QPushButton("Analyze and Parse")
        self.start_button.setProperty("role", "primary")
        self.start_button.clicked.connect(self.start_analysis)
        button_layout.addWidget(self.start_button)

        self.copy_button = QPushButton("Copy Task IDs to Clipboard")
        self.copy_button.setProperty("role", "primary")
        self.copy_button.clicked.connect(self.copy_to_clipboard)
        self.copy_button.setEnabled(False)
        button_layout.addWidget(self.copy_button)
        
        self.refresh_button = QPushButton("Refresh Files")
        self.refresh_button.clicked.connect(lambda: (self.clear_extra_files(), self.populate_downloads_files()))
        button_layout.addWidget(self.refresh_button)
        
        self.clear_button = QPushButton("Clear Output")
        self.clear_button.clicked.connect(self.clear_output)
        button_layout.addWidget(self.clear_button)

        self.log_toggle_button = QPushButton("Show Log")
        self.log_toggle_button.setCheckable(True)
        self.log_toggle_button.setToolTip(
            "Show or hide the raw activity log (technical details)."
        )
        self.log_toggle_button.toggled.connect(self._toggle_log)
        button_layout.addWidget(self.log_toggle_button)

        self.undo_button = QPushButton("Undo")
        self.undo_button.setToolTip(f"Undo last table edit  (Ctrl+Z) — remembers up to {UNDO_HISTORY_LIMIT} steps")
        self.undo_button.setShortcut("Ctrl+Z")
        self.undo_button.clicked.connect(self._undo_edit)
        self.undo_button.setEnabled(False)
        button_layout.addWidget(self.undo_button)

        self.redo_button = QPushButton("Redo")
        self.redo_button.setToolTip("Redo last undone edit  (Ctrl+Y)")
        self.redo_button.setShortcut("Ctrl+Y")
        self.redo_button.clicked.connect(self._redo_edit)
        self.redo_button.setEnabled(False)
        button_layout.addWidget(self.redo_button)


        
        main_layout.addLayout(button_layout)

        # Table widget — fills all remaining vertical space.
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(0)
        self.table_widget.setRowCount(0)
        self.table_widget.currentItemChanged.connect(self._on_current_item_changed)
        self.table_widget.itemChanged.connect(self._on_cell_text_changed)
        main_layout.addWidget(self.table_widget, 1)  # stretch=1

        # Log pane — hidden by default; revealed via the "Show Log" toggle button.
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        monospace_font = QFont("Courier New", 9)
        monospace_font.setFixedPitch(True)
        self.output_text.setFont(monospace_font)
        self.output_text.setFixedHeight(150)
        self.output_text.hide()
        main_layout.addWidget(self.output_text)
        
        # Set layout
        central_widget.setLayout(main_layout)
        self._apply_theme()

    def _apply_theme(self) -> None:
        """Apply the Cool Teal color scheme, adapting to the system light/dark theme."""
        _palette = QApplication.instance().palette()
        is_dark = _palette.color(_palette.ColorRole.Window).lightness() < 128

        if is_dark:
            qss = """
                QWidget#central_widget { background-color: #1A2A2E; }
                QPushButton {
                    background-color: #2E7D32; color: white;
                    border: none; padding: 5px 12px; border-radius: 4px;
                }
                QPushButton:hover { background-color: #1B5E20; }
                QPushButton:pressed { background-color: #0A3D0A; }
                QPushButton:disabled { background-color: #1A2E1E; color: #4A7A4E; }
                QPushButton:checked { background-color: #1B5E20; border: 2px solid #A5D6A7; }
                QPushButton[role="primary"] { background-color: #006978; }
                QPushButton[role="primary"]:hover { background-color: #004D5C; }
                QPushButton[role="primary"]:pressed { background-color: #003040; }
                QPushButton[role="primary"]:disabled { background-color: #0A2A30; color: #1A5A65; }
                QHeaderView::section {
                    background-color: #006978; color: white;
                    padding: 4px; border: 1px solid #004D5C;
                }
                QMenuBar::item:selected { background-color: #006978; color: white; }
                QMenu::item:selected { background-color: #2E7D32; color: white; }
            """
            title_color = "#80CBC4"
            label_color = "#A5D6A7"
            status_color = "#ffffff"
        else:
            qss = """
                QWidget#central_widget { background-color: #C8E6EA; }
                QTableWidget, QTextEdit { background-color: #ffffff; }
                QPushButton {
                    background-color: #4CAF70; color: white;
                    border: none; padding: 5px 12px; border-radius: 4px;
                }
                QPushButton:hover { background-color: #388E55; }
                QPushButton:pressed { background-color: #2E7D45; }
                QPushButton:disabled { background-color: #C8E6D0; color: #9E9E9E; }
                QPushButton:checked { background-color: #2E7D45; border: 2px solid #1B5E35; }
                QPushButton[role="primary"] { background-color: #1A97A4; }
                QPushButton[role="primary"]:hover { background-color: #147D88; }
                QPushButton[role="primary"]:pressed { background-color: #0D636D; }
                QPushButton[role="primary"]:disabled { background-color: #B2EBF2; color: #9E9E9E; }
                QHeaderView::section {
                    background-color: #1A97A4; color: white;
                    padding: 4px; border: 1px solid #147D88;
                }
                QMenuBar::item:selected { background-color: #1A97A4; color: white; }
                QMenu::item:selected { background-color: #4CAF70; color: white; }
            """
            title_color = "#00695C"
            label_color = "#00695C"
            status_color = "#000000"

        QApplication.instance().setStyleSheet(qss)
        self.title_label.setStyleSheet(f"QLabel {{ color: {title_color}; }}")
        self.status_label.setStyleSheet(
            f"QLabel {{ font-style: italic; font-weight: bold; color: {status_color}; padding: 3px 0; }}"
        )
        self.detected_label.setStyleSheet(f"QLabel {{ color: {label_color}; font-style: italic; }}")
        self.extra_files_label.setStyleSheet(f"QLabel {{ color: {label_color}; font-style: italic; }}")

    def populate_downloads_files(self):
        """Populate the file combo box with CSV and Excel files from Downloads"""
        self.file_combo.clear()
        downloads_path = Path.home() / "Downloads"
        
        files = []
        for ext in ['*.csv', '*.xlsx', '*.xls']:
            files.extend(downloads_path.glob(ext))
        
        if files:
            # Sort by modification time (most recent first)
            files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            
            for file in files:
                self.file_combo.addItem(file.name, str(file))
            
            self.status_label.setText(f"Found {len(files)} file(s) in Downloads")
        else:
            self.file_combo.addItem("No files found", None)
            self.status_label.setText("No CSV or Excel files found in Downloads folder")

    # ------------------------------------------------------------------
    # Multi-file helpers
    # ------------------------------------------------------------------

    def add_extra_files(self):
        """Open a file browser and append chosen files to the extra-files list."""
        # Refresh the primary file combo first so any newly-downloaded files
        # are already indexed before the user picks additional files.
        self.populate_downloads_files()
        downloads = str(Path.home() / "Downloads")
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Add files to parse",
            downloads,
            "Data files (*.csv *.xlsx *.xls);;All files (*)",
        )
        existing_paths = {
            self.extra_files_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self.extra_files_list.count())
        }
        for fp in files:
            if fp in existing_paths:
                continue
            item = QListWidgetItem(Path(fp).name)
            item.setData(Qt.ItemDataRole.UserRole, fp)
            self.extra_files_list.addItem(item)
            existing_paths.add(fp)
        self._refresh_extra_files_ui()

    def remove_extra_file(self):
        """Remove the currently selected item(s) from the extra-files list."""
        for item in self.extra_files_list.selectedItems():
            self.extra_files_list.takeItem(self.extra_files_list.row(item))
        self._refresh_extra_files_ui()

    def clear_extra_files(self):
        """Remove all additional files from the list."""
        self.extra_files_list.clear()
        self._refresh_extra_files_ui()

    def _refresh_extra_files_ui(self):
        """Sync the label text and list-widget visibility to the current count."""
        count = self.extra_files_list.count()
        if count == 0:
            self.extra_files_label.setText("Additional files: none")
            self.extra_files_list.hide()
        else:
            noun = "file" if count == 1 else "files"
            self.extra_files_label.setText(f"Additional files: {count} {noun} selected")
            self.extra_files_list.show()

    def start_analysis(self):
        """Start the analysis in a worker thread"""
        if self.file_combo.currentData() is None:
            QMessageBox.warning(self, "No File", "No file selected. Please select a file from Downloads.")
            return

        primary_filepath = self.file_combo.currentData()
        extra_filepaths = [
            self.extra_files_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self.extra_files_list.count())
        ]
        all_filepaths = [primary_filepath] + extra_filepaths
        selected_function = self.function_combo.currentData()

        # Prompt for sheet selection for every multi-sheet workbook in the list.
        sheet_names: Dict[str, str] = {}
        for fp in all_filepaths:
            sheet = self._pick_sheet(fp)
            if sheet is False:  # user cancelled
                return
            if sheet:
                sheet_names[fp] = sheet

        # Disable buttons while processing
        self.start_button.setEnabled(False)
        self.file_combo.setEnabled(False)
        self.function_combo.setEnabled(False)
        self.refresh_button.setEnabled(False)
        
        # Clear the previous table
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)
        self.strikethrough_rows.clear()
        self.task_ids = None
        self.copy_button.setEnabled(False)

        file_count = len(all_filepaths)
        if file_count > 1:
            self.status_label.setText(f"Analyzing {file_count} files...")
        else:
            self.status_label.setText("Analyzing...")

        # Create and start worker thread
        self.worker_thread = WorkerThread(all_filepaths, selected_function, self.registry, sheet_names=sheet_names)
        self.worker_thread.output.connect(self.append_output)
        self.worker_thread.error.connect(self.on_error)
        self.worker_thread.table_ready.connect(self.on_table_ready)
        self.worker_thread.task_ids_ready.connect(self.on_task_ids_ready)
        self.worker_thread.data_loaded.connect(self.on_data_loaded)
        self.worker_thread.template_matched.connect(self.on_template_matched)
        self.worker_thread.finished.connect(self.on_finished)
        self.worker_thread.start()
    
    def append_output(self, text: str):
        """Append text to the output display"""
        self.output_text.append(text)

    def on_template_matched(self, template_name: str, reason: str):
        """Update the 'Detected category' label after template selection."""
        self.detected_label.setText(f"Detected category: {template_name}")
        self.detected_label.setToolTip(f"Matched by: {reason}")

    def _is_dark_theme(self) -> bool:
        """Determine whether the active application theme is dark."""
        palette = self.table_widget.palette()
        window_color = palette.color(palette.ColorRole.Window)
        return window_color.lightness() < 128

    def _default_table_text_color(self) -> QColor:
        """Return standard table text color based on active theme."""
        return QColor("white") if self._is_dark_theme() else QColor("black")

    def _is_highlighted_cell(self, data_row_idx: int, col_name: str) -> bool:
        """Return True when a cell has an active background highlight."""
        return (data_row_idx, col_name) in self.cell_highlights

    @staticmethod
    def _color_for_name(name: str) -> Optional[QColor]:
        """Map template highlight colour names to QColor swatches."""
        return {
            "darkgreen": QColor(130, 200, 150),
            "darkyellow": QColor(230, 200, 90),
            "red": QColor(220, 120, 120),
            "blue": QColor(140, 180, 230),
        }.get(name)

    def on_table_ready(self, df: pd.DataFrame, meta: ViewMeta):
        """Populate the table with data and add the checkbox column."""
        if df is None or df.empty:
            self.output_text.append("No data to display in table.\n")
            return

        self._suppressing_history = True

        # Remember the view-shaped df + highlights so "Export view..." can
        # write exactly what the user is looking at.
        self.view_df = df.copy()
        self.view_meta = meta
        self.export_action.setEnabled(True)

        self.strikethrough_rows.clear()

        # Build cell-highlight lookup from the template-driven ViewMeta.
        self.cell_highlights = {(int(r), c): color for (r, c, color) in meta.highlights}

        # Use the template-detected location column for divider rows when present.
        location_col = meta.location_column if meta.location_column in df.columns else None
        self.location_col = location_col

        render_rows = []
        self.table_row_map = []
        self.table_row_location_values = []
        prev_prefix = None
        for df_idx, (_, row_data) in enumerate(df.iterrows()):
            prefix = None
            location_text = None
            if location_col:
                location_value = row_data.get(location_col)
                location_text = "" if pd.isna(location_value) else str(location_value)
                parsed_prefix, parsed_number, _ = parse_location_parts(location_text)
                prefix = (parsed_prefix, parsed_number // 100000 if parsed_number != float("inf") else float("inf"))

            if location_col and prev_prefix is not None and prefix != prev_prefix:
                render_rows.append({"type": "divider"})
                self.table_row_map.append(None)
                self.table_row_location_values.append(None)

            render_rows.append({"type": "data", "df_idx": df_idx, "row_data": row_data})
            self.table_row_map.append(df_idx)
            self.table_row_location_values.append(location_text if location_col else None)
            if location_col:
                prev_prefix = prefix

        self.table_widget.setRowCount(len(render_rows))
        self.table_widget.setColumnCount(len(df.columns) + 1)

        headers = list(df.columns) + ["Done?"]
        self.table_widget.setHorizontalHeaderLabels(headers)

        default_text_color = self._default_table_text_color()

        for row_idx, render_row in enumerate(render_rows):
            if render_row["type"] == "divider":
                for col_idx in range(len(df.columns)):
                    item = QTableWidgetItem("-----")
                    item.setFlags(Qt.ItemFlag.NoItemFlags)
                    item.setForeground(default_text_color)
                    self.table_widget.setItem(row_idx, col_idx, item)
                continue

            row_data = render_row["row_data"]
            df_idx = render_row["df_idx"]

            for col_idx, col_name in enumerate(df.columns):
                value = row_data[col_name]
                cell_text = "" if pd.isna(value) else str(value)
                item = QTableWidgetItem(cell_text)
                color_name = self.cell_highlights.get((df_idx, col_name))
                if color_name:
                    qcolor = self._color_for_name(color_name)
                    if qcolor is not None:
                        item.setBackground(qcolor)
                        item.setForeground(QColor("black"))
                self.table_widget.setItem(row_idx, col_idx, item)

            checkbox = QCheckBox()
            checkbox.stateChanged.connect(lambda checked, r=row_idx: self.on_checkbox_changed(r, checked))
            self.table_widget.setCellWidget(row_idx, len(df.columns), checkbox)

        self._suppressing_history = False
        self._edit_history.clear()
        self._update_undo_redo_buttons()
        self.table_widget.resizeColumnsToContents()
        header = self.table_widget.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(True)
    
    def on_checkbox_changed(self, row_idx: int, state):
        """Handle checkbox state change — apply/remove strikethrough and record for undo."""
        if self.bulk_checkbox_update or self._suppressing_history:
            return

        is_checked = state == 2  # Qt.CheckState.Checked is 2

        data_row_idx = None
        if isinstance(self.table_row_map, list) and row_idx < len(self.table_row_map):
            data_row_idx = self.table_row_map[row_idx]
        if data_row_idx is None:
            return

        location_value = None
        if isinstance(self.table_row_location_values, list) and row_idx < len(self.table_row_location_values):
            location_value = self.table_row_location_values[row_idx]

        self._suppressing_history = True
        try:
            if location_value is None:
                # Single-row edit — checkbox already flipped to is_checked.
                affected = [(row_idx, not is_checked, is_checked)]
                self.apply_row_strikethrough(row_idx, is_checked)
            else:
                # Location group — bulk-update every row sharing this location.
                affected = []
                self.bulk_checkbox_update = True
                try:
                    for tidx, row_loc in enumerate(self.table_row_location_values):
                        if row_loc != location_value:
                            continue
                        checkbox = self.table_widget.cellWidget(tidx, self.table_widget.columnCount() - 1)
                        if isinstance(checkbox, QCheckBox):
                            # Primary row already flipped; others haven't changed yet.
                            old = (not is_checked) if tidx == row_idx else checkbox.isChecked()
                            affected.append((tidx, old, is_checked))
                            if checkbox.isChecked() != is_checked:
                                checkbox.setChecked(is_checked)
                        self.apply_row_strikethrough(tidx, is_checked)
                finally:
                    self.bulk_checkbox_update = False
        finally:
            self._suppressing_history = False

        self._edit_history.push(_CheckboxEdit(affected))
        self._update_undo_redo_buttons()

    def _on_current_item_changed(self, current, previous):
        """Cache the current cell text when focus moves, so undo can restore it."""
        if current is not None and not self._suppressing_history:
            self._pre_edit_text[(current.row(), current.column())] = current.text()

    def _on_cell_text_changed(self, item):
        """Record a user-driven cell text change to the undo history."""
        if self._suppressing_history:
            return
        key = (item.row(), item.column())
        old_text = self._pre_edit_text.get(key)
        if old_text is not None and item.text() != old_text:
            self._edit_history.push(_CellTextEdit(
                row=item.row(), col=item.column(),
                old_text=old_text, new_text=item.text(),
            ))
            self._pre_edit_text[key] = item.text()
            self._update_undo_redo_buttons()

    def _undo_edit(self):
        """Undo the most recent table edit."""
        edit = self._edit_history.undo()
        if edit is None:
            return
        self._suppressing_history = True
        self.bulk_checkbox_update = True
        try:
            if isinstance(edit, _CheckboxEdit):
                for tidx, old_checked, _new in edit.affected:
                    cb = self.table_widget.cellWidget(tidx, self.table_widget.columnCount() - 1)
                    if isinstance(cb, QCheckBox):
                        cb.setChecked(old_checked)
                    self.apply_row_strikethrough(tidx, old_checked)
            elif isinstance(edit, _CellTextEdit):
                item = self.table_widget.item(edit.row, edit.col)
                if item:
                    item.setText(edit.old_text)
                    self._pre_edit_text[(edit.row, edit.col)] = edit.old_text
        finally:
            self._suppressing_history = False
            self.bulk_checkbox_update = False
        self._update_undo_redo_buttons()

    def _redo_edit(self):
        """Reapply the most recently undone table edit."""
        edit = self._edit_history.redo()
        if edit is None:
            return
        self._suppressing_history = True
        self.bulk_checkbox_update = True
        try:
            if isinstance(edit, _CheckboxEdit):
                for tidx, _old, new_checked in edit.affected:
                    cb = self.table_widget.cellWidget(tidx, self.table_widget.columnCount() - 1)
                    if isinstance(cb, QCheckBox):
                        cb.setChecked(new_checked)
                    self.apply_row_strikethrough(tidx, new_checked)
            elif isinstance(edit, _CellTextEdit):
                item = self.table_widget.item(edit.row, edit.col)
                if item:
                    item.setText(edit.new_text)
                    self._pre_edit_text[(edit.row, edit.col)] = edit.new_text
        finally:
            self._suppressing_history = False
            self.bulk_checkbox_update = False
        self._update_undo_redo_buttons()

    def _update_undo_redo_buttons(self):
        """Sync Undo / Redo button enabled state to the history stacks."""
        self.undo_button.setEnabled(self._edit_history.can_undo())
        self.redo_button.setEnabled(self._edit_history.can_redo())

    def apply_row_strikethrough(self, row_idx: int, is_checked: bool):
        """Apply or remove strikethrough for a table row."""
        data_row_idx = None
        if isinstance(self.table_row_map, list) and row_idx < len(self.table_row_map):
            data_row_idx = self.table_row_map[row_idx]
        if data_row_idx is None:
            return
        
        if is_checked and row_idx not in self.strikethrough_rows:
            # Apply strikethrough and red color
            self.strikethrough_rows.add(row_idx)
            for col_idx in range(self.table_widget.columnCount() - 1):  # Skip checkbox column
                item = self.table_widget.item(row_idx, col_idx)
                if item:
                    font = item.font()
                    font.setStrikeOut(True)
                    item.setFont(font)
                    item.setForeground(QColor("red"))
        elif not is_checked and row_idx in self.strikethrough_rows:
            # Remove strikethrough and restore original formatting
            self.strikethrough_rows.discard(row_idx)
            default_text_color = self._default_table_text_color()
            
            for col_idx in range(self.table_widget.columnCount() - 1):  # Skip checkbox column
                item = self.table_widget.item(row_idx, col_idx)
                if item:
                    font = item.font()
                    font.setStrikeOut(False)
                    item.setFont(font)
                    
                    # Get column name
                    header_item = self.table_widget.horizontalHeaderItem(col_idx)
                    if header_item is None:
                        continue
                    col_name = header_item.text()
                    
                    # Restore text color: black for highlighted cells, theme-default for others
                    if self._is_highlighted_cell(data_row_idx, col_name):
                        item.setForeground(QColor("black"))
                    else:
                        item.setForeground(default_text_color)
    
    def on_error(self, error_msg: str):
        """Handle errors from worker thread"""
        self.output_text.append(f"\nERROR: {error_msg}\n")
        self.status_label.setText("Error occurred")
    
    def on_task_ids_ready(self, task_ids: list):
        """Handle task IDs from worker thread"""
        self.task_ids = task_ids
        self.copy_button.setEnabled(bool(task_ids))
    
    def on_data_loaded(self, df: pd.DataFrame):
        """Handle data loaded from worker thread"""
        self.current_df = df
        self.export_action.setEnabled(df is not None and not df.empty)
    
    def on_finished(self):
        """Called when worker thread finishes"""
        # Re-enable buttons
        self.start_button.setEnabled(True)
        self.file_combo.setEnabled(True)
        self.function_combo.setEnabled(True)
        self.refresh_button.setEnabled(True)
        
        self.status_label.setText("Analysis complete")
    
    
    def clear_output(self):
        """Clear the output display and table"""
        self.output_text.clear()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)
        self.strikethrough_rows.clear()
        self.task_ids = None
        self.copy_button.setEnabled(False)
        self._edit_history.clear()
        self._pre_edit_text.clear()
        self._update_undo_redo_buttons()
        self.status_label.setText("Output cleared")

    def _toggle_log(self, checked: bool):
        """Show or hide the raw activity log pane."""
        if checked:
            self.output_text.show()
            self.log_toggle_button.setText("Hide Log")
        else:
            self.output_text.hide()
            self.log_toggle_button.setText("Show Log")

    def open_templates_dialog(self):
        """Show the templates CRUD dialog and reload the registry on close."""
        dialog = TemplatesDialog(self.registry, self)
        dialog.exec()
        # Re-load from disk so any external edits are picked up too.
        self.registry = TemplateRegistry.load()

    def show_version_info(self):
        """Display the current and previously installed version."""
        previous = self.settings.value("version/previous", "", type=str)
        if previous:
            message = (
                f"<b>Current version:</b> {__version__}<br><br>"
                f"<b>Previous version:</b> {previous}"
            )
        else:
            message = (
                f"<b>Current version:</b> {__version__}<br><br>"
                "<i>No previous version recorded.</i>"
            )
        QMessageBox.information(self, "Version Info", message)

    def _pick_sheet(self, filepath: str):
        """Resolve which sheet to read for an Excel file.

        Returns:
            - The sheet name (str) when one is chosen,
            - "" when the file is a CSV or has at most one sheet (caller will
              pass ``None`` to ``WorkerThread`` and let pandas default),
            - ``False`` when the user cancels the dialog (caller should abort).
        """
        if filepath.lower().endswith(".csv"):
            return ""
        try:
            sheets = ExcelParser(filepath).list_sheets()
        except Exception:
            return ""
        if len(sheets) <= 1:
            return ""

        key = f"sheets/{filepath}"
        last = self.settings.value(key, type=str) or sheets[0]
        try:
            current_index = sheets.index(last)
        except ValueError:
            current_index = 0

        chosen, ok = QInputDialog.getItem(
            self,
            "Select sheet",
            f"This workbook contains {len(sheets)} sheets. Pick one:",
            sheets,
            current_index,
            False,
        )
        if not ok:
            return False
        self.settings.setValue(key, chosen)
        return chosen

    def export_view(self):
        """Export the current view (template-applied dataframe) to CSV or XLSX.

        For .xlsx, cell highlights from the active ViewMeta are preserved
        using openpyxl pattern fills.
        """
        if self.view_df is None or self.view_df.empty:
            QMessageBox.information(self, "Nothing to export", "Run an analysis first.")
            return

        default_name = "docureader_view.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export current view",
            default_name,
            "Excel Workbook (*.xlsx);;CSV (*.csv)",
        )
        if not path:
            return

        try:
            if path.lower().endswith(".csv"):
                self.view_df.to_csv(path, index=False)
            else:
                self._export_xlsx_with_highlights(path)
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return

        self.status_label.setText(f"Exported view to {path}")
        self.output_text.append(f"\nExported view to {path}\n")

    def _export_xlsx_with_highlights(self, path: str) -> None:
        """Write ``self.view_df`` to ``path`` and apply highlight fills."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        df = self.view_df
        meta = self.view_meta
        wb = Workbook()
        ws = wb.active
        ws.title = "View"

        cols = list(df.columns)
        ws.append(cols)
        for _, row in df.iterrows():
            ws.append(["" if pd.isna(v) else v for v in row.tolist()])

        # Map our colour names to hex fills.
        hex_for = {
            "darkgreen": "82C896",
            "darkyellow": "E6C85A",
            "red": "DC7878",
            "blue": "8CB4E6",
        }
        if meta is not None and meta.highlights:
            col_index = {c: i + 1 for i, c in enumerate(cols)}  # openpyxl is 1-based
            for (data_row, col_name, color_name) in meta.highlights:
                hex_code = hex_for.get(color_name)
                col_idx = col_index.get(col_name)
                if not hex_code or col_idx is None:
                    continue
                # +2: header row offset (row 1) + 0-based data_row.
                ws.cell(row=int(data_row) + 2, column=col_idx).fill = PatternFill(
                    start_color=hex_code, end_color=hex_code, fill_type="solid"
                )

        wb.save(path)

    def batch_export(self):
        """Pick N source files, apply each one's matched template, and write
        one ``<source>.view.xlsx`` per file into a chosen output folder.

        Runs synchronously on the GUI thread - intended for short batches
        from Downloads. Highlights are preserved (same export path as the
        single-file export).
        """
        downloads = str(Path.home() / "Downloads")
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select files to batch process",
            downloads,
            "Data files (*.csv *.xlsx *.xls);;All files (*)",
        )
        if not files:
            return

        out_dir = QFileDialog.getExistingDirectory(
            self, "Select output folder", downloads
        )
        if not out_dir:
            return

        out_path = Path(out_dir)
        successes = 0
        failures: list[tuple[str, str]] = []

        self.output_text.append("\n" + "=" * 60 + "\n")
        self.output_text.append(f"Batch processing {len(files)} file(s) -> {out_dir}\n")
        self.output_text.append("=" * 60 + "\n")

        prev_view_df = self.view_df
        prev_view_meta = self.view_meta

        for src in files:
            src_path = Path(src)
            try:
                parser = ExcelParser(str(src_path))
                parser.read_excel()  # pandas defaults; sheet selector is interactive only.
                if parser.df is None:
                    raise RuntimeError("failed to load")

                match = self.registry.select(parser.df.columns, str(src_path))
                view_df, meta = apply_template(parser.df, match.template)

                # Reuse the existing exporter by stashing the view temporarily.
                self.view_df = view_df
                self.view_meta = meta
                dst = out_path / f"{src_path.stem}.view.xlsx"
                self._export_xlsx_with_highlights(str(dst))

                self.output_text.append(
                    f"  OK  {src_path.name}  -> {dst.name}  "
                    f"[template: {match.template.name}, {len(view_df)} rows]\n"
                )
                successes += 1
            except Exception as e:
                failures.append((src_path.name, str(e)))
                self.output_text.append(f"  FAIL  {src_path.name}: {e}\n")

        # Restore the on-screen view's state.
        self.view_df = prev_view_df
        self.view_meta = prev_view_meta

        self.output_text.append(
            f"\nBatch complete: {successes} succeeded, {len(failures)} failed.\n"
        )
        self.status_label.setText(
            f"Batch export: {successes}/{len(files)} succeeded"
        )
        if failures:
            QMessageBox.warning(
                self,
                "Batch export finished with errors",
                f"{successes} of {len(files)} files exported.\n\n"
                + "\n".join(f"- {n}: {e}" for n, e in failures[:10])
                + ("" if len(failures) <= 10 else f"\n... and {len(failures) - 10} more"),
            )

    def copy_to_clipboard(self):
        """Copy task IDs to clipboard"""
        if self.task_ids:
            clipboard_text = ", ".join(map(str, self.task_ids))
            clipboard = QApplication.clipboard()
            if clipboard is not None:
                clipboard.setText(clipboard_text)
            self.status_label.setText(f"Copied {len(self.task_ids)} Task ID(s) to clipboard")
            self.output_text.append(f"\nCopied to clipboard: {clipboard_text}\n")

    def check_for_updates(self):
        """Download and stage application updates, showing a live progress dialog."""
        reply = QMessageBox.question(
            self,
            "Update Application",
            "This will check GitHub for a newer release and download it if available.\n\n"
            "The application will close and relaunch automatically after the update is applied.\n\n"
            "Do you want to continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        self.output_text.append("\n" + "=" * 60 + "\n")
        self.output_text.append("Checking for updates…\n")
        self.update_action.setEnabled(False)
        self.update_action.setText("Updating…")

        thread = _UpdateThread(include_prereleases=self.prerelease_action.isChecked())
        dialog = _UpdateProgressDialog(self)
        self._update_thread = thread

        # Closures capture _result so the main flow can read them after exec().
        _result: dict = {}

        def on_staged(staged_dir, install_dir):
            _result["staged_dir"] = staged_dir
            _result["install_dir"] = install_dir
            dialog.accept()

        def on_error(msg: str):
            _result["error"] = msg
            dialog.reject()

        thread.progress.connect(dialog.set_progress)
        thread.status.connect(dialog.set_status)
        thread.staged.connect(on_staged)
        thread.error.connect(on_error)

        thread.start()
        accepted = dialog.exec() == QDialog.DialogCode.Accepted

        # Regardless of outcome, tell the thread to stop and wait for it.
        thread.cancel()
        thread.wait(8000)
        self._update_thread = None

        self.update_action.setEnabled(True)
        self.update_action.setText("Check && Install Updates")

        if not accepted:
            err = _result.get("error", "")
            if err:
                QMessageBox.warning(self, "Update", err)
                self.output_text.append(f"[Update] {err}\n")
            else:
                self.output_text.append("[Update] Cancelled.\n")
            self.status_label.setText("Update cancelled")
            return

        staged_dir = _result.get("staged_dir")
        install_dir = _result.get("install_dir")
        if staged_dir is None or install_dir is None:
            return

        self.output_text.append(f"[Update] Staged at: {staged_dir}\n")
        self.status_label.setText("Update downloaded — restart to apply")

        restart = QMessageBox.question(
            self,
            "Update Ready",
            "The update has been downloaded and verified.\n\n"
            "Click Yes to close the application now and apply the update.\n"
            "DocuReader will relaunch automatically once the installation is complete.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if restart == QMessageBox.StandardButton.Yes:
            import updater_github
            rc = updater_github.apply_update(Path(staged_dir), Path(install_dir))
            if rc != 0:
                QMessageBox.critical(
                    self,
                    "Update Failed",
                    "Failed to launch the update installer.\n\n"
                    "Please try again or reinstall manually.",
                )
                self.status_label.setText("Update apply failed.")
                return
            self.close()
        else:
            self.status_label.setText("Update staged — restart when ready.")

    def terminate_program(self):
        """Terminate the application"""
        reply = QMessageBox.question(
            self,
            "Terminate Program",
            "Are you sure you want to close the application?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Stop worker thread if running
            if self.worker_thread and self.worker_thread.isRunning():
                self.worker_thread.quit()
                self.worker_thread.wait()
            
            self.close()
    
    def closeEvent(self, event):
        """Handle window close event"""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.quit()
            self.worker_thread.wait()
        if self._update_thread and self._update_thread.isRunning():
            self._update_thread.cancel()
            self._update_thread.wait(3000)
        event.accept()


def main():
    # First-run migration: if launched from %ProgramFiles% (the legacy admin
    # install location), copy ourselves to %LOCALAPPDATA% and relaunch so
    # subsequent auto-updates can write without UAC.
    try:
        from migrate import maybe_migrate_install
        if maybe_migrate_install():
            return
    except Exception:
        pass

    app = QApplication(sys.argv)
    gui = ExcelParserGUI()
    gui.show()
    # Centre on the primary screen after the window frame is established.
    screen_geo = QApplication.primaryScreen().availableGeometry()
    fg = gui.frameGeometry()
    fg.moveCenter(screen_geo.center())
    gui.move(fg.topLeft())
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
